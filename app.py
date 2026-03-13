"""
app.py — PDF Belge → Excel Otomasyon
Dil desteği: Türkçe / Русский
"""
import sys, threading, datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
sys.path.insert(0, str(Path(__file__).parent))
from core.parser import parse_pdf
from core.excel_writer import process_letter_to_excel, execute_write_operations

LANG = {
"tr": {
"app_title":"PDF Belge → Excel Otomasyon","header_title":"PDF → Excel","lang_btn":"RU",
"ready":"Hazır — PDF dosyalarını yükleyin",
"tab_files":"  1 · Dosyalar  ","tab_preview":"  2 · Önizleme  ","tab_write":"  3 · Excel Yazma  ","tab_log":"  4 · Log  ",
"pdf_section":"  PDF Belgeler  ","btn_add_pdf":"+ PDF Ekle","btn_clear":"Temizle",
"pdf_hint":"Dağıtım mektubu + УПД dosyalarını seçin",
"col_filename":"Dosya adı","col_type":"Tip","col_status":"Durum",
"excel_section":"  Excel Dosyası  ","not_selected":"— seçilmedi —","btn_select_excel":"Excel Seç",
"sheet_label":"Sayfa adı:","btn_process":"⚙  PDF Dosyalarını İşle →",
"validation_sec":"  Doğrulama  ","not_processed":"PDF işlenmedi.","extracted_sec":"  Çıkarılan Veriler  ",
"btn_match":"🔗  Excel ile Eşleştir →",
"write_plan_sec":"  Yazma Planı  ","col_row":"Satır","col_col":"Sütun","col_header":"Başlık","col_value":"Değer","col_desc":"Açıklama",
"dry_run_chk":"Test modu (dry-run) — Excel'e yazmadan önizle",
"btn_preview":"👁  Önizle (Dry-Run)","btn_write":"✍  Excel'e Yaz",
"log_label":"İşlem günlüğü","btn_clear_log":"Günlüğü Temizle",
"warn_no_pdf":"Önce PDF dosyası ekleyin.","warn_no_letter":"Önce PDF'leri işleyin (Sekme 1).",
"warn_no_excel":"Excel dosyası seçilmedi (Sekme 1).","warn_no_match":"Önce eşleştirme yapın (Sekme 2).",
"confirm_write":"Excel dosyasına yazılacak!\n\nOrijinal dosya değiştirilmez — '_updated' uzantılı yeni dosya oluşturulur.\n\nDevam edilsin mi?",
"confirm_title":"Onay","success_title":"Başarılı","success_msg":"Excel dosyası güncellendi!\n\n","error_title":"Hata","match_error":"Eşleştirme Hatası",
"lbl_number":"Numara","lbl_amount":"Tutar","lbl_supplier":"Tedarikçi","lbl_inn":"INN",
"lbl_invoice":"Счёт №","lbl_contract":"Sözleşme","lbl_spec":"Спец. №","lbl_upds":"УПД'ler",
"lbl_smeta":"Smeta pos.","lbl_buyer":"Alıcı","lbl_total":"Toplam","lbl_materials":"Malzemeler",
"no_letter":"⚠  Dağıtım mektubu bulunamadı",
"val_ok":"✓  УПД toplamı = mektup tutarı",
"val_diff":"⚠  Fark: {diff:,.2f} ₽  (mektup: {letter:.2f}, УПД: {upd:.2f})",
"val_no_letter":"⚠  Dağıtım mektubu yüklenmedi",
"processing":"PDF'ler işleniyor...","matching":"Eşleştiriliyor...",
"writing_dry":"Test yazma...","writing":"Excel'e yazılıyor...",
"status_letter_ok":"✓ Mektup: bulundu","status_letter_no":"⚠ Mektup: YOK",
"status_upd":"{n} УПД","ops_planned":"✓ {n} işlem planlandı · {col}",
"preview_done":"Önizleme tamamlandı","write_done":"Yazma tamamlandı","msgs":"{n} mesaj",
"type_letter":"Dağ. Mektubu","type_upd":"УПД","type_invoice":"Счёт","type_unknown":"?",
"status_ok":"✓ OK","status_pending":"Bekliyor","processing_row":"İşleniyor",
},
"ru": {
"app_title":"PDF Документ → Excel Автоматизация","header_title":"PDF → Excel","lang_btn":"TR",
"ready":"Готово — загрузите PDF файлы",
"tab_files":"  1 · Файлы  ","tab_preview":"  2 · Предпросмотр  ","tab_write":"  3 · Запись в Excel  ","tab_log":"  4 · Журнал  ",
"pdf_section":"  PDF Документы  ","btn_add_pdf":"+ Добавить PDF","btn_clear":"Очистить",
"pdf_hint":"Выберите распред. письмо + УПД файлы",
"col_filename":"Имя файла","col_type":"Тип","col_status":"Статус",
"excel_section":"  Excel Файл  ","not_selected":"— не выбран —","btn_select_excel":"Выбрать Excel",
"sheet_label":"Лист:","btn_process":"⚙  Обработать PDF файлы →",
"validation_sec":"  Проверка  ","not_processed":"PDF не обработан.","extracted_sec":"  Извлечённые данные  ",
"btn_match":"🔗  Сопоставить с Excel →",
"write_plan_sec":"  План записи  ","col_row":"Строка","col_col":"Столбец","col_header":"Заголовок","col_value":"Значение","col_desc":"Описание",
"dry_run_chk":"Тест режим (dry-run) — предпросмотр без записи",
"btn_preview":"👁  Предпросмотр (Dry-Run)","btn_write":"✍  Записать в Excel",
"log_label":"Журнал операций","btn_clear_log":"Очистить журнал",
"warn_no_pdf":"Сначала добавьте PDF файлы.","warn_no_letter":"Сначала обработайте PDF (Вкладка 1).",
"warn_no_excel":"Excel файл не выбран (Вкладка 1).","warn_no_match":"Сначала выполните сопоставление (Вкладка 2).",
"confirm_write":"Будет выполнена запись в Excel!\n\nОригинальный файл не изменится — будет создан новый файл с суффиксом '_updated'.\n\nПродолжить?",
"confirm_title":"Подтверждение","success_title":"Успешно","success_msg":"Excel файл обновлён!\n\n","error_title":"Ошибка","match_error":"Ошибка сопоставления",
"lbl_number":"Номер","lbl_amount":"Сумма","lbl_supplier":"Поставщик","lbl_inn":"ИНН",
"lbl_invoice":"Счёт №","lbl_contract":"Договор","lbl_spec":"Спец. №","lbl_upds":"УПД",
"lbl_smeta":"Позиция сметы","lbl_buyer":"Покупатель","lbl_total":"Итого","lbl_materials":"Материалы",
"no_letter":"⚠  Распределительное письмо не найдено",
"val_ok":"✓  Сумма УПД = сумма письма",
"val_diff":"⚠  Расхождение: {diff:,.2f} ₽  (письмо: {letter:.2f}, УПД: {upd:.2f})",
"val_no_letter":"⚠  Распределительное письмо не загружено",
"processing":"Обработка PDF...","matching":"Сопоставление...",
"writing_dry":"Тестовая запись...","writing":"Запись в Excel...",
"status_letter_ok":"✓ Письмо: найдено","status_letter_no":"⚠ Письмо: НЕТ",
"status_upd":"{n} УПД","ops_planned":"✓ {n} операций · {col}",
"preview_done":"Предпросмотр завершён","write_done":"Запись завершена","msgs":"{n} сообщений",
"type_letter":"Распред. письмо","type_upd":"УПД","type_invoice":"Счёт","type_unknown":"?",
"status_ok":"✓ OK","status_pending":"Ожидание","processing_row":"Обработка",
},
}

BG_MAIN="#F8F7F5"; BG_CARD="#FFFFFF"; BG_ACCENT="#1A1A1A"
FG_MAIN="#1A1A1A"; FG_MUTED="#6B6B6B"; FG_OK="#2D7D46"; FG_WARN="#B45309"; FG_ERR="#B91C1C"
BORDER="#E2E0D8"
FONT_SM=("Segoe UI",9); FONT_MD=("Segoe UI",10); FONT_BOLD=("Segoe UI",10,"bold")
FONT_HEAD=("Segoe UI",13,"bold"); FONT_MONO=("Consolas",9); PAD=12

class AppState:
    def __init__(self):
        self.pdf_files=[]; self.excel_path=""; self.sheet_name=""
        self.parse_results=[]; self.letter=None; self.upd_list=[]; self.match_result=None
        self.dry_run=True; self.lang="tr"

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.state=AppState()
        self.geometry("920x700"); self.minsize(780,560); self.configure(bg=BG_MAIN)
        self._spinner_frames=["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
        self._spinner_idx=0
        self._spinner_job=None
        self._spinning_rows=set()
        self._setup_styles(); self._build_ui(); self._apply_lang()

    def T(self,key): return LANG[self.state.lang].get(key,key)

    def _toggle_lang(self):
        self.state.lang="ru" if self.state.lang=="tr" else "tr"
        self._apply_lang()

    def _apply_lang(self):
        t=self.T
        self.title(t("app_title")); self.header_lbl.configure(text=t("header_title"))
        self.lang_btn.configure(text=t("lang_btn"))
        self.nb.tab(0,text=t("tab_files")); self.nb.tab(1,text=t("tab_preview"))
        self.nb.tab(2,text=t("tab_write")); self.nb.tab(3,text=t("tab_log"))
        self.pdf_frame.configure(text=t("pdf_section"))
        self.btn_add.configure(text=t("btn_add_pdf")); self.btn_clr.configure(text=t("btn_clear"))
        self.pdf_hint_lbl.configure(text=t("pdf_hint"))
        self.pdf_tree.heading("fn",text=t("col_filename")); self.pdf_tree.heading("tp",text=t("col_type")); self.pdf_tree.heading("st",text=t("col_status"))
        self.excel_frame.configure(text=t("excel_section"))
        if not self.state.excel_path: self.excel_path_var.set(t("not_selected"))
        self.btn_excel.configure(text=t("btn_select_excel"))
        self.sheet_lbl.configure(text=t("sheet_label")); self.btn_process.configure(text=t("btn_process"))
        self.validation_frame.configure(text=t("validation_sec"))
        self.detail_frame.configure(text=t("extracted_sec")); self.btn_match.configure(text=t("btn_match"))
        self.write_plan_frame.configure(text=t("write_plan_sec"))
        self.plan_tree.heading("r",text=t("col_row")); self.plan_tree.heading("c",text=t("col_col"))
        self.plan_tree.heading("h",text=t("col_header")); self.plan_tree.heading("v",text=t("col_value")); self.plan_tree.heading("d",text=t("col_desc"))
        self.dry_run_chk.configure(text=t("dry_run_chk"))
        self.btn_prev_wr.configure(text=t("btn_preview")); self.btn_wr.configure(text=t("btn_write"))
        self.log_lbl.configure(text=t("log_label")); self.btn_clr_log.configure(text=t("btn_clear_log"))
        self._update_status(t("ready"),"info")

    def _setup_styles(self):
        s=ttk.Style(self); s.theme_use("clam")
        s.configure("TFrame",background=BG_MAIN)
        s.configure("TLabel",background=BG_MAIN,foreground=FG_MAIN,font=FONT_MD)
        s.configure("Muted.TLabel",background=BG_MAIN,foreground=FG_MUTED,font=FONT_SM)
        s.configure("Head.TLabel",background=BG_MAIN,foreground=FG_MAIN,font=FONT_HEAD)
        s.configure("TButton",font=FONT_MD,padding=(10,5))
        s.configure("Primary.TButton",background=BG_ACCENT,foreground="white",font=FONT_BOLD,relief="flat")
        s.map("Primary.TButton",background=[("active","#333"),("disabled","#CCC")])
        s.configure("TNotebook",background=BG_MAIN,tabmargins=[2,4,0,0])
        s.configure("TNotebook.Tab",font=FONT_MD,padding=(12,6))
        s.configure("Treeview",background=BG_CARD,fieldbackground=BG_CARD,font=FONT_MD,rowheight=26)
        s.configure("Treeview.Heading",font=FONT_BOLD,background=BG_MAIN,foreground=FG_MUTED)
        s.configure("TProgressbar",troughcolor=BORDER,background=BG_ACCENT,thickness=4)

    def _build_ui(self):
        hdr=ttk.Frame(self,style="TFrame",padding=(PAD,10,PAD,6)); hdr.pack(fill="x")
        self.header_lbl=ttk.Label(hdr,text="",style="Head.TLabel"); self.header_lbl.pack(side="left")
        self.status_lbl=ttk.Label(hdr,text="",font=FONT_SM,background=BG_MAIN); self.status_lbl.pack(side="right",padx=(0,8))
        self.lang_btn=ttk.Button(hdr,text="",command=self._toggle_lang); self.lang_btn.pack(side="right",padx=(0,4))
        ttk.Separator(self,orient="horizontal").pack(fill="x")
        self.nb=ttk.Notebook(self); self.nb.pack(fill="both",expand=True,padx=PAD,pady=PAD)
        self.tab_files=ttk.Frame(self.nb,style="TFrame",padding=PAD)
        self.tab_preview=ttk.Frame(self.nb,style="TFrame",padding=PAD)
        self.tab_write=ttk.Frame(self.nb,style="TFrame",padding=PAD)
        self.tab_log=ttk.Frame(self.nb,style="TFrame",padding=PAD)
        for tab in [self.tab_files,self.tab_preview,self.tab_write,self.tab_log]:
            self.nb.add(tab,text="")
        self._build_tab1(); self._build_tab2(); self._build_tab3(); self._build_tab4()

    def _build_tab1(self):
        f=self.tab_files
        self.pdf_frame=ttk.LabelFrame(f,text="",padding=(PAD,8)); self.pdf_frame.pack(fill="x",pady=(0,PAD))
        br=ttk.Frame(self.pdf_frame); br.pack(fill="x",pady=(0,8))
        self.btn_add=ttk.Button(br,text="",command=self._add_pdfs); self.btn_add.pack(side="left",padx=(0,6))
        self.btn_clr=ttk.Button(br,text="",command=self._clear_pdfs); self.btn_clr.pack(side="left")
        self.pdf_hint_lbl=ttk.Label(br,text="",style="Muted.TLabel"); self.pdf_hint_lbl.pack(side="right")
        lf=ttk.Frame(self.pdf_frame); lf.pack(fill="x")
        self.pdf_tree=ttk.Treeview(lf,columns=("fn","tp","st"),show="headings",height=5)
        for c,w in [("fn",380),("tp",130),("st",160)]: self.pdf_tree.column(c,width=w)
        sb=ttk.Scrollbar(lf,orient="vertical",command=self.pdf_tree.yview)
        self.pdf_tree.configure(yscrollcommand=sb.set)
        self.pdf_tree.pack(side="left",fill="x",expand=True); sb.pack(side="right",fill="y")
        self.excel_frame=ttk.LabelFrame(f,text="",padding=(PAD,8)); self.excel_frame.pack(fill="x",pady=(0,PAD))
        xr=ttk.Frame(self.excel_frame); xr.pack(fill="x",pady=(0,6))
        self.excel_path_var=tk.StringVar()
        ttk.Label(xr,textvariable=self.excel_path_var,foreground=FG_MUTED,font=FONT_SM,background=BG_MAIN).pack(side="left",fill="x",expand=True)
        self.btn_excel=ttk.Button(xr,text="",command=self._select_excel); self.btn_excel.pack(side="right")
        sr=ttk.Frame(self.excel_frame); sr.pack(fill="x")
        self.sheet_lbl=ttk.Label(sr,text=""); self.sheet_lbl.pack(side="left",padx=(0,6))
        self.sheet_var=tk.StringVar(value=self.state.sheet_name)
        self.sheet_combo=ttk.Combobox(sr,textvariable=self.sheet_var,width=28,state="normal")
        self.sheet_combo.pack(side="left")
        self.sheet_hint=ttk.Label(sr,text="",style="Muted.TLabel"); self.sheet_hint.pack(side="left",padx=(8,0))
        self.sheet_var.trace_add("write",lambda *_:setattr(self.state,"sheet_name",self.sheet_var.get()))
        ttk.Separator(f,orient="horizontal").pack(fill="x",pady=PAD)
        self.btn_process=ttk.Button(f,text="",style="Primary.TButton",command=self._run_parse); self.btn_process.pack(anchor="e")
        self.parse_progress=ttk.Progressbar(f,mode="indeterminate",length=200); self.parse_progress.pack(anchor="e",pady=(6,0))

    def _build_tab2(self):
        f=self.tab_preview
        self.validation_frame=ttk.LabelFrame(f,text="",padding=(PAD,8)); self.validation_frame.pack(fill="x",pady=(0,PAD))
        self.validation_lbl=ttk.Label(self.validation_frame,text="",foreground=FG_MUTED,background=BG_MAIN,font=FONT_SM); self.validation_lbl.pack(anchor="w")
        self.detail_frame=ttk.LabelFrame(f,text="",padding=(PAD,8)); self.detail_frame.pack(fill="both",expand=True)
        self.detail_text=scrolledtext.ScrolledText(self.detail_frame,font=FONT_MONO,bg=BG_CARD,fg=FG_MAIN,relief="flat",wrap="word",height=16)
        self.detail_text.pack(fill="both",expand=True); self.detail_text.configure(state="disabled")
        ttk.Separator(f,orient="horizontal").pack(fill="x",pady=PAD)
        self.btn_match=ttk.Button(f,text="",style="Primary.TButton",command=self._run_match); self.btn_match.pack(anchor="e")

    def _build_tab3(self):
        f=self.tab_write
        self.write_plan_frame=ttk.LabelFrame(f,text="",padding=(PAD,8)); self.write_plan_frame.pack(fill="both",expand=True,pady=(0,PAD))
        self.plan_tree=ttk.Treeview(self.write_plan_frame,columns=("r","c","h","v","d"),show="headings",height=8)
        for col,w in [("r",60),("c",70),("h",200),("v",140),("d",260)]: self.plan_tree.column(col,width=w)
        self.plan_tree.column("r",anchor="center"); self.plan_tree.column("c",anchor="center"); self.plan_tree.column("v",anchor="e")
        sb=ttk.Scrollbar(self.write_plan_frame,orient="vertical",command=self.plan_tree.yview)
        self.plan_tree.configure(yscrollcommand=sb.set)
        self.plan_tree.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")
        of=ttk.Frame(f); of.pack(fill="x",pady=(0,PAD))
        self.dry_run_var=tk.BooleanVar(value=True)
        self.dry_run_chk=ttk.Checkbutton(of,text="",variable=self.dry_run_var,command=lambda:setattr(self.state,"dry_run",self.dry_run_var.get()))
        self.dry_run_chk.pack(side="left")
        self.write_status=ttk.Label(of,text="",background=BG_MAIN,font=FONT_SM); self.write_status.pack(side="right")
        ttk.Separator(f,orient="horizontal").pack(fill="x",pady=(0,PAD))
        br=ttk.Frame(f); br.pack(fill="x")
        self.btn_prev_wr=ttk.Button(br,text="",command=lambda:self._run_write(True)); self.btn_prev_wr.pack(side="left",padx=(0,8))
        self.btn_wr=ttk.Button(br,text="",style="Primary.TButton",command=lambda:self._run_write(False)); self.btn_wr.pack(side="right")
        self.write_progress=ttk.Progressbar(f,mode="indeterminate",length=200); self.write_progress.pack(anchor="e",pady=(6,0))

    def _build_tab4(self):
        f=self.tab_log
        self.log_lbl=ttk.Label(f,text="",style="Muted.TLabel"); self.log_lbl.pack(anchor="w",pady=(0,6))
        self.log_text=scrolledtext.ScrolledText(f,font=FONT_MONO,bg=BG_CARD,fg=FG_MAIN,relief="flat",wrap="word",height=22)
        self.log_text.pack(fill="both",expand=True)
        self.btn_clr_log=ttk.Button(f,text="",command=lambda:self.log_text.delete("1.0","end")); self.btn_clr_log.pack(anchor="e",pady=(6,0))

    def _add_pdfs(self):
        files=filedialog.askopenfilenames(title=self.T("btn_add_pdf"),filetypes=[("PDF","*.pdf *.PDF"),("*","*.*")])
        for fp in files:
            if fp not in self.state.pdf_files:
                self.state.pdf_files.append(fp)
                self.pdf_tree.insert("","end",values=(Path(fp).name,"—",self.T("status_pending")))
        self._log(f"+{len(files)} PDF")

    def _load_sheet_names(self,path):
        try:
            import openpyxl
            wb=openpyxl.load_workbook(path,read_only=True)
            sheets=wb.sheetnames; wb.close()
            self.sheet_combo['values']=sheets
            if sheets:
                # İlk 'Дог.' ile başlayan sayfayı seç, yoksa ilkini
                default=next((s for s in sheets if 'дог' in s.lower()),sheets[0])
                self.sheet_var.set(default)
                hint=f"({len(sheets)} sayfa)"
                self.sheet_hint.configure(text=hint)
        except Exception as e:
            self._log(f"Sayfa yüklenemedi: {e}")

    def _clear_pdfs(self):
        self.state.pdf_files.clear(); self.state.parse_results.clear()
        self.state.letter=None; self.state.upd_list.clear()
        for iid in self.pdf_tree.get_children(): self.pdf_tree.delete(iid)

    def _select_excel(self):
        path=filedialog.askopenfilename(title=self.T("btn_select_excel"),filetypes=[("Excel","*.xlsx *.xlsm"),("*","*.*")])
        if path:
            self.state.excel_path=path; self.excel_path_var.set(Path(path).name); self._log(f"Excel: {path}")
            self._load_sheet_names(path)

    def _run_parse(self):
        if not self.state.pdf_files: messagebox.showwarning("",self.T("warn_no_pdf")); return
        self.parse_progress.start(10); self._update_status(self.T("processing"),"info")
        self._spinning_rows=set(range(len(self.pdf_tree.get_children())))
        self._spinner_idx=0
        self._tick_spinner()
        threading.Thread(target=self._parse_worker,daemon=True).start()

    def _parse_worker(self):
        letters,upds,results=[],[],[]
        try:
            for i,fp in enumerate(self.state.pdf_files):
                self.after(0,self._set_active_row,i)
                res=parse_pdf(fp); results.append(res)
                if res.doc_type=="letter" and res.letter: letters.append(res.letter)
                elif res.doc_type=="upd" and res.upd: upds.append(res.upd)
                self._spinning_rows.discard(i)
                self.after(0,self._update_tree_row,i,res)
        except Exception as e:
            import traceback
            self.after(0,self._log,f'PARSE EXCEPTION: {traceback.format_exc()}')
        self.state.parse_results=results; self.state.letter=letters[0] if letters else None; self.state.upd_list=upds
        self.after(0,self._parse_done)


    def _tick_spinner(self):
        if not self._spinning_rows and self._spinner_job:
            self._stop_spinner(); return
        frame=self._spinner_frames[self._spinner_idx % len(self._spinner_frames)]
        self._spinner_idx+=1
        children=self.pdf_tree.get_children()
        for i in self._spinning_rows:
            if i<len(children):
                vals=list(self.pdf_tree.item(children[i],"values"))
                if len(vals)>=3:
                    label=self.T("processing_row")
                    vals[2]=f"{frame} {label}..."
                    self.pdf_tree.item(children[i],values=vals)
        self._spinner_job=self.after(120,self._tick_spinner)

    def _stop_spinner(self):
        if self._spinner_job:
            self.after_cancel(self._spinner_job); self._spinner_job=None
        self._spinning_rows=set()

    def _set_active_row(self,index):
        children=self.pdf_tree.get_children()
        if index>=len(children): return
        vals=list(self.pdf_tree.item(children[index],"values"))
        if len(vals)>=3: vals[2]=f"⏳ {self.T('processing_row')}..."; self.pdf_tree.item(children[index],values=vals)

    def _update_tree_row(self,index,res):
        children=self.pdf_tree.get_children()
        if index>=len(children): return
        tm={"letter":self.T("type_letter"),"upd":self.T("type_upd"),"invoice":self.T("type_invoice"),"spec":"Спец.","contract":"Договор","unknown":self.T("type_unknown")}
        status=self.T("status_ok") if not res.errors else f"⚠ {res.errors[0][:28]}"
        self.pdf_tree.item(children[index],values=(Path(res.file_path).name,tm.get(res.doc_type,res.doc_type),status))

    def _parse_done(self):
        self._stop_spinner()
        self.parse_progress.stop(); self._update_preview(); self.nb.select(1)
        ok=self.state.letter is not None
        self._update_status((self.T("status_letter_ok") if ok else self.T("status_letter_no"))+" · "+self.T("status_upd").format(n=len(self.state.upd_list)),"ok" if ok else "warn")

    def _update_preview(self):
        self.detail_text.configure(state="normal"); self.detail_text.delete("1.0","end"); t=self.T
        if self.state.letter:
            L=self.state.letter
            self.detail_text.insert("end","══════════════════════════════════\n")
            for lbl,val in [(t("lbl_number"),f"исх.№{L.number} от {L.date}"),(t("lbl_amount"),f"{L.amount:,.2f} ₽"),
                (t("lbl_supplier"),L.supplier_name),(t("lbl_inn"),L.supplier_inn),(t("lbl_invoice"),f"{L.invoice_number} от {L.invoice_date}"),
                (t("lbl_contract"),L.contract_number),(t("lbl_spec"),L.specification_number),
                (t("lbl_upds"),", ".join(L.upd_numbers)),(t("lbl_smeta"),L.smeta_position)]:
                self.detail_text.insert("end",f"  {lbl:<16}: {val}\n")
        else:
            self.detail_text.insert("end",t("no_letter")+"\n")
        for upd in self.state.upd_list:
            self.detail_text.insert("end",f"\n── УПД №{upd.number} ({upd.date}) ──\n")
            for lbl,val in [(t("lbl_supplier"),upd.supplier_name),(t("lbl_buyer"),upd.buyer_name),(t("lbl_total"),f"{upd.total_incl_vat:,.2f} ₽")]:
                self.detail_text.insert("end",f"  {lbl:<16}: {val}\n")
            if upd.items:
                self.detail_text.insert("end",f"  {t('lbl_materials')}:\n")
                for item in upd.items:
                    self.detail_text.insert("end",f"    • {item.material_name[:52]}\n      {item.quantity} {item.unit} × {item.price_excl_vat:,.2f} ₽ (НДС {int(item.vat_rate*100)}%) = {item.total_incl_vat:,.2f} ₽\n")
        if self.state.letter and self.state.upd_list:
            ut=sum(u.total_incl_vat for u in self.state.upd_list); la=self.state.letter.amount; diff=abs(ut-la)
            vt=t("val_ok") if diff<1 else t("val_diff").format(diff=diff,letter=la,upd=ut)
            vc=FG_OK if diff<1 else FG_WARN
        elif not self.state.letter: vt=t("val_no_letter"); vc=FG_WARN
        else: vt="—"; vc=FG_MUTED
        self.validation_lbl.configure(text=vt,foreground=vc)
        self.detail_text.configure(state="disabled")

    def _run_match(self):
        if not self.state.letter: messagebox.showwarning("",self.T("warn_no_letter")); return
        if not self.state.excel_path: messagebox.showwarning("",self.T("warn_no_excel")); return
        self._update_status(self.T("matching"),"info")
        threading.Thread(target=self._match_worker,daemon=True).start()

    def _match_worker(self):
        try:
            result=process_letter_to_excel(letter=self.state.letter,upd_list=self.state.upd_list,
                excel_path=self.state.excel_path,sheet_name=self.state.sheet_name,dry_run=True)
            self.state.match_result=result
        except Exception as e:
            import traceback
            err=traceback.format_exc()
            from core.excel_writer import MatchResult
            self.state.match_result=MatchResult(supplier_col=None,target_rows=[],write_ops=[],errors=[str(e)],warnings=[])
            self.after(0,self._log,f'EXCEPTION: {err}')
        self.after(0,self._match_done)

    def _match_done(self):
        result=self.state.match_result
        if result.errors:
            self._update_status(f"✗ {result.errors[0][:50]}","error"); self._log("ERROR: "+" | ".join(result.errors))
            messagebox.showerror(self.T("match_error"),"\n".join(result.errors)); return
        for iid in self.plan_tree.get_children(): self.plan_tree.delete(iid)
        from openpyxl.utils import get_column_letter
        for op in result.write_ops:
            self.plan_tree.insert("","end",values=(op.row,get_column_letter(op.col),op.description[:30],
                f"{op.value:,.2f}" if isinstance(op.value,float) else str(op.value),op.description))
        ci=(f"Col {result.supplier_col.col_start} ({result.supplier_col.invoice_key})" if result.supplier_col else "?")
        self._update_status(self.T("ops_planned").format(n=len(result.write_ops),col=ci),"ok")
        self._log(f"Match OK: {len(result.write_ops)} ops · {ci}")
        if result.warnings: self._log("WARN: "+" | ".join(result.warnings))
        self.nb.select(2)

    def _run_write(self,dry_run=True):
        if not self.state.match_result: messagebox.showwarning("",self.T("warn_no_match")); return
        if not dry_run and not messagebox.askyesno(self.T("confirm_title"),self.T("confirm_write")): return
        self.write_progress.start(10)
        self._update_status(self.T("writing_dry") if dry_run else self.T("writing"),"info")
        threading.Thread(target=self._write_worker,args=(dry_run,),daemon=True).start()

    def _write_worker(self,dry_run):
        try:
            success,messages=execute_write_operations(excel_path=self.state.excel_path,sheet_name=self.state.sheet_name,
                ops=self.state.match_result.write_ops,supplier_col=self.state.match_result.supplier_col,dry_run=dry_run)
        except Exception as e:
            import traceback
            success,messages=False,[f'EXCEPTION: {traceback.format_exc()}']
        self.after(0,self._write_done,success,messages,dry_run)

    def _write_done(self,success,messages,dry_run):
        self.write_progress.stop()
        for m in messages: self._log(("[DRY] " if dry_run else "")+m)
        label=self.T("preview_done") if dry_run else self.T("write_done")
        if success:
            s=f"✓ {label} · {self.T('msgs').format(n=len(messages))}"
            self._update_status(s,"ok"); self.write_status.configure(text=s,foreground=FG_OK)
            if not dry_run:
                messagebox.showinfo(self.T("success_title"),self.T("success_msg")+"\n".join(messages[-3:])); self.nb.select(3)
        else:
            self._update_status("✗","error"); self.write_status.configure(text="✗",foreground=FG_ERR)
            messagebox.showerror(self.T("error_title"),"\n".join(messages))

    def _update_status(self,text,level="info"):
        self.status_lbl.configure(text=text,foreground={"info":FG_MUTED,"ok":FG_OK,"warn":FG_WARN,"error":FG_ERR}.get(level,FG_MUTED))

    def _log(self,text):
        ts=datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end",f"[{ts}] {text}\n"); self.log_text.see("end")

if __name__=="__main__":
    app=MainApp(); app.mainloop()
