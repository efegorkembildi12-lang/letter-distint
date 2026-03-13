"""
excel_writer.py — LFT Excel dosyasına veri eşleştirme ve yazma
Hedef dosya: ЛФТ_260218_Договор__95721.xlsx  (ve benzerleri)
Hedef sayfa: "Дог.95721"

Sütun yapısı (her tedarikçi = 4 sütun bloku):
  Sütun N+0: Объем (teslimat miktarı)
  Sütun N+1: Цена (birim fiyat)
  Sütun N+2: Стоимость по договору поставки
  Sütun N+3: Оплата

Row 2:  "нет в смете" özet satırı (smeta'da olmayan malzemeler)
Row 6:  Tedarikçi başlık satırı (счёт + УПД numaraları)
Row 7:  Sütun başlıkları (Объем / Цена / Стоимость / Оплата)
Row 9+: Smeta malzeme satırları
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Union

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from core.parser import DistributionLetter, UPDDocument


# ─── Veri modelleri ────────────────────────────────────────────────────────────

@dataclass
class SupplierColumn:
    """Tespit edilen tedarikçi sütun grubu"""
    invoice_key: str      # Arama anahtarı (örn: "670951")
    header_text: str      # Row 6 hücre metni
    col_start: int        # Объем sütunu (1-indexed)
    col_price: int        # Цена
    col_amount: int       # Стоимость
    col_payment: int      # Оплата
    row_header: int = 6   # Başlık satırı


@dataclass
class TargetRow:
    """Yazılacak hedef satır"""
    row_number: int
    row_type: str         # "net_v_smete" | "smeta_item"
    item_name: str        # Satır adı/açıklaması
    confidence: float     # 0.0–1.0


@dataclass
class WriteOperation:
    """Tek bir yazma işlemi"""
    row: int
    col: int
    value: Union[float, str]
    description: str      # Kullanıcıya gösterilecek açıklama


@dataclass
class MatchResult:
    """Eşleştirme sonucu"""
    supplier_col: Optional[SupplierColumn]
    target_rows: List[TargetRow]
    write_ops: List[WriteOperation]
    errors: List[str]
    warnings: List[str]


# ─── Sütun tespiti ─────────────────────────────────────────────────────────────

def find_supplier_column(ws, invoice_key: str) -> Optional[SupplierColumn]:
    """
    Row 6'yı tarayarak fatura numarasıyla eşleşen tedarikçi sütununu bul.
    invoice_key: örn "670951" (счёт numarasının ayırt edici kısmı)
    """
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]

    for col_idx, cell_val in enumerate(row6):
        if cell_val and isinstance(cell_val, str):
            if invoice_key.lower() in cell_val.lower():
                start = col_idx + 1  # 1-indexed
                return SupplierColumn(
                    invoice_key=invoice_key,
                    header_text=cell_val,
                    col_start=start,        # Объем
                    col_price=start + 1,    # Цена
                    col_amount=start + 2,   # Стоимость
                    col_payment=start + 3,  # Оплата
                )
    return None


def extract_invoice_key(invoice_number: str) -> str:
    """
    "00ЦБ-670951" → "670951"
    "969/2504067-1" → "2504067"
    Sayısal kısmı al (prefix/suffix olmadan)
    """
    # Son sayısal bloğu al
    nums = re.findall(r'\d+', invoice_number)
    if not nums:
        return invoice_number
    # En uzun sayısal bloğu döndür
    return max(nums, key=len)


# ─── Satır tespiti ─────────────────────────────────────────────────────────────

def find_target_rows(ws, supplier_col: SupplierColumn,
                     material_names: List[str]) -> List[TargetRow]:
    """
    Malzemelerin yazılacağı satırları tespit et.
    Önce smeta satırlarında ara; bulunamazsa Row 2 (нет в смете) öner.
    """
    targets = []
    max_row = ws.max_row

    for mat_name in material_names:
        best_row = None
        best_score = 0.0

        # Smeta satırlarında benzerlik ara (sütun 7 = Наименование)
        for row_idx in range(9, min(max_row + 1, 600)):
            cell_name = ws.cell(row_idx, 7).value
            if cell_name and isinstance(cell_name, str):
                score = _name_similarity(mat_name, cell_name)
                if score > best_score and score > 0.4:
                    best_score = score
                    best_row = TargetRow(
                        row_number=row_idx,
                        row_type="smeta_item",
                        item_name=cell_name,
                        confidence=score,
                    )

        if best_row and best_score >= 0.6:
            targets.append(best_row)
        else:
            # Smeta'da yok → Row 2 нет в смете
            targets.append(TargetRow(
                row_number=2,
                row_type="net_v_smete",
                item_name=mat_name,
                confidence=1.0,
            ))

    # Deduplikasyon: aynı satıra birden fazla malzeme düşebilir
    seen_rows = {}
    deduped = []
    for t in targets:
        if t.row_number not in seen_rows:
            seen_rows[t.row_number] = t
            deduped.append(t)

    return deduped


def _name_similarity(query: str, candidate: str) -> float:
    """
    Basit token-tabanlı benzerlik skoru.
    Kablo adı için optimize: marka, kesit, voltaj bilgilerini karşılaştır.
    """
    def normalize(s: str) -> set[str]:
        s = s.lower()
        # Önemli token'ları çıkar: harf+rakam kombinasyonları
        tokens = re.findall(r'[а-яёa-z]+|\d+[,.]?\d*', s)
        return set(tokens)

    q_tokens = normalize(query)
    c_tokens = normalize(candidate)
    if not q_tokens or not c_tokens:
        return 0.0
    intersection = q_tokens & c_tokens
    union = q_tokens | c_tokens
    return len(intersection) / len(union)


# ─── Yazma işlemleri planlama ──────────────────────────────────────────────────

def plan_write_operations(
    letter: DistributionLetter,
    upd_list: List[UPDDocument],
    supplier_col: SupplierColumn,
    target_rows: List[TargetRow],
) -> List[WriteOperation]:
    """
    Tüm verileri inceleyerek hangi hücrelere ne yazılacağını planla.
    Henüz Excel'e yazmaz — önizleme için.
    """
    ops = []

    # Row 2 нет в смете satırı için toplam tutar
    net_v_smete_rows = [t for t in target_rows if t.row_type == "net_v_smete"]
    if net_v_smete_rows:
        # Mektup tutarını Row 2 Стоимость hücresine
        ops.append(WriteOperation(
            row=2,
            col=supplier_col.col_amount,
            value=letter.amount,
            description=f"Нет в смете · Стоимость = {letter.amount:,.2f} ₽"
        ))

    # Smeta satırları için УПД verilerini yaz
    for upd in upd_list:
        for item in upd.items:
            # Bu malzeme için hedef satır bul
            best_target = None
            best_score = 0.0
            for t in target_rows:
                if t.row_type == "smeta_item":
                    score = _name_similarity(item.material_name, t.item_name)
                    if score > best_score:
                        best_score = score
                        best_target = t

            if best_target and best_score >= 0.5:
                # Объем
                ops.append(WriteOperation(
                    row=best_target.row_number,
                    col=supplier_col.col_start,
                    value=item.quantity,
                    description=f"Объем: {item.quantity} {item.unit}"
                ))
                # Цена (без НДС)
                ops.append(WriteOperation(
                    row=best_target.row_number,
                    col=supplier_col.col_price,
                    value=item.price_excl_vat,
                    description=f"Цена без НДС: {item.price_excl_vat:,.2f} ₽"
                ))
                # Стоимость (с НДС)
                ops.append(WriteOperation(
                    row=best_target.row_number,
                    col=supplier_col.col_amount,
                    value=item.total_incl_vat,
                    description=f"Стоимость с НДС: {item.total_incl_vat:,.2f} ₽"
                ))

    return ops


# ─── Excel'e yazma ─────────────────────────────────────────────────────────────

HIGHLIGHT_FILL = None  # openpyxl kuruluysa aşağıda init edilir

def _init_styles():
    global HIGHLIGHT_FILL
    if openpyxl and HIGHLIGHT_FILL is None:
        HIGHLIGHT_FILL = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
        )


def execute_write_operations(
    excel_path: str,
    sheet_name: str,
    ops: List[WriteOperation],
    hyperlinks:Optional[ Dict[str, str] ] = None,
    supplier_col: Optional[SupplierColumn] = None,
    dry_run: bool = False,
) -> Tuple[bool, List[str]]:
    """
    Planlanan yazma işlemlerini Excel dosyasına uygula.

    Args:
        excel_path:   LFT dosya yolu
        sheet_name:   Hedef sayfa adı (örn: "Дог.95721")
        ops:          plan_write_operations'dan gelen liste
        hyperlinks:   {belge_adı: dosya_yolu} → Row 6 başlığa yorum olarak eklenir
        supplier_col: Başlık hücresine hyperlink eklemek için
        dry_run:      True ise yazma yapma, sadece doğrula

    Returns:
        (başarı_bool, mesaj_listesi)
    """
    if openpyxl is None:
        return False, ["openpyxl kurulu değil. pip install openpyxl"]

    _init_styles()
    messages = []

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        return False, [f"Excel dosyası açılamadı: {e}"]

    if sheet_name not in wb.sheetnames:
        return False, [f"Sayfa bulunamadı: '{sheet_name}'. Mevcut sayfalar: {wb.sheetnames}"]

    ws = wb[sheet_name]

    if dry_run:
        messages.append(f"[DRY RUN] {len(ops)} yazma işlemi planlandı:")
        for op in ops:
            col_letter = get_column_letter(op.col)
            messages.append(f"  {col_letter}{op.row} ← {op.value}  ({op.description})")
        return True, messages

    # Yazma işlemlerini uygula
    written = 0
    for op in ops:
        cell = ws.cell(op.row, op.col)

        # Mevcut değer varsa uyar (sıfır değil)
        existing = cell.value
        if existing and existing not in (0, "0", "", None):
            messages.append(
                f"⚠️  {get_column_letter(op.col)}{op.row}: "
                f"Mevcut değer ({existing}) üzerine yazılıyor → {op.value}"
            )

        cell.value = op.value

        # Yeni yazılan hücreyi hafif yeşil ile işaretle
        if HIGHLIGHT_FILL:
            cell.fill = HIGHLIGHT_FILL

        written += 1

    # Hyperlink yorumları ekle (Row 6 başlık hücresine)
    if hyperlinks and supplier_col:
        header_cell = ws.cell(supplier_col.row_header, supplier_col.col_start)
        existing_text = header_cell.value or ""
        link_text = "\n".join([f"→ {name}: {path}" for name, path in hyperlinks.items()])
        # Excel yorumu olarak ekle (hyperlink desteği sınırlı)
        from openpyxl.comments import Comment
        comment = Comment(
            text=f"Bağlantılı belgeler:\n{link_text}",
            author="LFT Otomasyon"
        )
        header_cell.comment = comment
        messages.append(f"✓ {len(hyperlinks)} belge referansı Row 6 yorumuna eklendi")

    # Kaydet — orijinal bozulmasın diye yeni isimle
    out_path = Path(excel_path)
    backup_path = out_path.with_stem(out_path.stem + "_updated")
    try:
        wb.save(str(backup_path))
        messages.append(f"✓ {written} hücre yazıldı → {backup_path.name}")
        return True, messages
    except Exception as e:
        return False, [f"Dosya kaydedilemedi: {e}"]


# ─── Yüksek seviye: tam iş akışı ──────────────────────────────────────────────

def process_letter_to_excel(
    letter: DistributionLetter,
    upd_list: List[UPDDocument],
    excel_path: str,
    sheet_name: str = "Дог.95721",
    hyperlinks:Optional[ Dict[str, str] ] = None,
    dry_run: bool = False,
) -> MatchResult:
    """
    Dağıtım mektubu + УПД verilerini LFT Excel'e işle.
    Tüm adımları kapsar: tespit → planlama → yazma.
    """
    if openpyxl is None:
        return MatchResult(
            supplier_col=None, target_rows=[], write_ops=[],
            errors=["openpyxl kurulu değil"], warnings=[]
        )

    errors, warnings = [], []

    # Excel aç
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
    except Exception as e:
        return MatchResult(
            supplier_col=None, target_rows=[], write_ops=[],
            errors=[str(e)], warnings=[]
        )

    # 1. Tedarikçi sütununu bul
    inv_key = extract_invoice_key(letter.invoice_number)
    supplier_col = find_supplier_column(ws, inv_key)
    if not supplier_col:
        errors.append(
            f"Tedarikçi sütunu bulunamadı. "
            f"Aranan anahtar: '{inv_key}' (fatura: {letter.invoice_number})"
        )
        return MatchResult(
            supplier_col=None, target_rows=[], write_ops=[],
            errors=errors, warnings=warnings
        )

    # 2. Malzeme adlarını topla (mektup + УПД'lerden)
    material_names = []
    for upd in upd_list:
        for item in upd.items:
            material_names.append(item.material_name)
    if not material_names:
        # Sadece toplam tutar yazılacak → нет в смете
        material_names = [f"[Toplam] {letter.invoice_number}"]
        warnings.append("УПД satır detayı bulunamadı — sadece toplam tutarı yazılacak")

    # 3. Hedef satırları bul
    target_rows = find_target_rows(ws, supplier_col, material_names)

    # 4. Yazma planı oluştur
    write_ops = plan_write_operations(letter, upd_list, supplier_col, target_rows)

    # 5. Yazma işlemlerini uygula
    success, msgs = execute_write_operations(
        excel_path=excel_path,
        sheet_name=sheet_name,
        ops=write_ops,
        hyperlinks=hyperlinks,
        supplier_col=supplier_col,
        dry_run=dry_run,
    )
    if not success:
        errors.extend(msgs)
    else:
        warnings.extend([m for m in msgs if m.startswith("⚠️")])

    return MatchResult(
        supplier_col=supplier_col,
        target_rows=target_rows,
        write_ops=write_ops,
        errors=errors,
        warnings=warnings,
    )
